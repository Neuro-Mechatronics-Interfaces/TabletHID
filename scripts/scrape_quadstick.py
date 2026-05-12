#!/usr/bin/env python3
"""
scrape_quadstick.py – Scrape QuadStick .xlsx config sheets into SQLite.

Usage
-----
    python scripts/scrape_quadstick.py [--root PATH] [--db PATH] [--csv-dir PATH]

Defaults
--------
    --root     G:\\Shared drives\\NML_shared\\Equipment Manuals and Software
               \\Quadstick\\Configurations
    --db       scripts/output/quadstick.db
    --csv-dir  scripts/output/  (set to empty string "" to skip CSV export)

Output tables
-------------
    configs       One row per mode-block (a "configuration" the user can switch to).
    bindings      One row per output->input mapping within a mode-block.
    inputs_ref    Canonical input names + descriptions (from "Inputs" sheets).
    outputs_ref   Canonical output names + descriptions (from "Outputs" sheets).
    preferences   Per-file device preference values (from "Preferences" sheets).

Useful views created automatically
-----------------------------------
    output_usage    Which outputs appear in how many configs (descending).
    input_usage     Which inputs are used how often (descending).
    config_summary  One row per config with boolean feature flags for clustering.

Sheet classification
--------------------
    Config sheet   Row 3 col A = "Output or Function"
    Inputs ref     Sheet named "Inputs"  OR  col A header = "Input Name"
    Outputs ref    Sheet named "Outputs" OR  col A header = "Output Name"
    Preferences    Sheet named "Preferences" OR col A = "Preference"
    (everything else is skipped)

Config sheet structure
----------------------
    Row 1  col A = "Profile Name"   col C, K, S, … = mode names (one per block)
    Row 2  col A = profile .csv     col C, K, S, … = layout type (Normal / etc.)
    Row 3  col A = "Output or Function"  col B = "Function"
           col C, K, S, … = transport per block (usb / bluetooth / …)
    Rows 4+  col A = output_name, col B = function_type,
             cols C–J = up to 8 inputs for block 0,
             cols K–R = up to 8 inputs for block 1, …
"""

import argparse
import csv
import sqlite3
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required.  Run:  pip install openpyxl")

DEFAULT_ROOT = (
    r"G:\Shared drives\NML_shared\Equipment Manuals and Software"
    r"\Quadstick\Configurations"
)
DEFAULT_DB = Path(__file__).parent / "output" / "quadstick.db"
DEFAULT_CSV_DIR = Path(__file__).parent / "output"

# ── helpers ────────────────────────────────────────────────────────────────────

def _nv(v) -> str | None:
    """Normalise a cell value to stripped string, or None if blank."""
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _load_rows(ws) -> list[list]:
    """Read all rows of a worksheet into a list of normalised-value lists."""
    return [[_nv(cell) for cell in row] for row in ws.iter_rows(values_only=True)]


def _get(row: list, col: int):
    """Safe column access."""
    return row[col] if col < len(row) else None

# ── sheet classification ───────────────────────────────────────────────────────

def classify_sheet(sheet_name: str, rows: list[list]) -> str:
    """Return one of: 'config', 'inputs_ref', 'outputs_ref', 'preferences', 'skip'."""
    sname = sheet_name.strip().lower()

    if sname == "inputs":
        return "inputs_ref"
    if sname == "outputs":
        return "outputs_ref"
    if sname in ("preferences", "preference"):
        return "preferences"

    if not rows:
        return "skip"

    # Config: row 3 col A = "Output or Function"
    if len(rows) >= 3 and _get(rows[2], 0) == "Output or Function":
        return "config"

    # Fallback: detect by first non-empty col-A header value
    for row in rows[:6]:
        v = _get(row, 0)
        if v == "Input Name":
            return "inputs_ref"
        if v == "Output Name":
            return "outputs_ref"
        if v in ("Preference", "Preferences"):
            return "preferences"

    return "skip"

# ── config sheet parser ────────────────────────────────────────────────────────

def parse_config_sheet(rows: list[list], xlsx_name: str, folder: str,
                        rel_path: str, sheet_name: str, sheet_index: int) -> tuple:
    """
    Parse a config sheet.

    Returns (configs, bindings):
        configs  – list of dicts, one per mode-block
        bindings – list of dicts, one per non-empty output row × mode-block,
                   keyed by block_index for later config_id assignment
    """
    if len(rows) < 4:
        return [], []

    r1, r2, r3 = rows[0], rows[1], rows[2]
    profile_file = _get(r2, 0)  # e.g. "android.csv"

    # Find mode-block start columns: positions ≥ 2 in row 3 that are non-None.
    # Each block is 8 columns wide (transport + 7 additional input slots).
    block_starts: list[int] = []
    r3_len = max(len(r3), len(r1) if r1 else 0)
    for ci in range(2, r3_len):
        if _get(r3, ci) is not None:
            block_starts.append(ci)

    if not block_starts:
        block_starts = [2]  # graceful fallback

    configs = []
    for block_idx, start_col in enumerate(block_starts):
        configs.append({
            "folder":       folder,
            "xlsx_name":    xlsx_name,
            "rel_path":     rel_path,
            "sheet_name":   sheet_name,
            "sheet_index":  sheet_index,
            "block_index":  block_idx,
            "mode_name":    _get(r1, start_col),
            "layout_type":  _get(r2, start_col),
            "transport":    _get(r3, start_col),
            "profile_file": profile_file,
        })

    bindings = []
    for row in rows[3:]:
        output_name = _get(row, 0)
        if output_name is None:
            continue
        function_type = _get(row, 1)

        for block_idx, start_col in enumerate(block_starts):
            inputs = [_get(row, start_col + offset) for offset in range(8)]
            if all(v is None for v in inputs):
                continue  # output not bound in this mode-block

            bindings.append({
                "_block_index": block_idx,
                "output_name":   output_name,
                "function_type": function_type,
                "input_1":  inputs[0],
                "input_2":  inputs[1],
                "input_3":  inputs[2],
                "input_4":  inputs[3],
                "input_5":  inputs[4],
                "input_6":  inputs[5],
                "input_7":  inputs[6],
                "input_8":  inputs[7],
            })

    return configs, bindings

# ── reference sheet parsers ────────────────────────────────────────────────────

def parse_inputs_ref(rows: list[list], xlsx_name: str) -> list[dict]:
    header_row = next(
        (i for i, r in enumerate(rows[:10]) if _get(r, 0) == "Input Name"), None
    )
    if header_row is None:
        return []
    results = []
    for row in rows[header_row + 1:]:
        name = _get(row, 0)
        if name is None:
            continue
        results.append({
            "input_name":  name,
            "description": _get(row, 1),
            "delay":       _get(row, 2),
            "source_xlsx": xlsx_name,
        })
    return results


def parse_outputs_ref(rows: list[list], xlsx_name: str) -> list[dict]:
    header_row = next(
        (i for i, r in enumerate(rows[:10]) if _get(r, 0) == "Output Name"), None
    )
    if header_row is None:
        return []
    results = []
    for row in rows[header_row + 1:]:
        name = _get(row, 0)
        if name is None:
            continue
        results.append({
            "output_name": name,
            "description": _get(row, 1),
            "source_xlsx": xlsx_name,
        })
    return results


def parse_preferences(rows: list[list], xlsx_name: str, sheet_name: str) -> list[dict]:
    header_row = next(
        (i for i, r in enumerate(rows[:10]) if _get(r, 0) == "Preference"), None
    )
    if header_row is None:
        return []
    results = []
    for row in rows[header_row + 1:]:
        name = _get(row, 0)
        if name is None:
            continue
        results.append({
            "source_xlsx": xlsx_name,
            "sheet_name":  sheet_name,
            "preference":  name,
            "value":       _get(row, 1),
            "units":       _get(row, 2),
            "description": _get(row, 3),
        })
    return results

# ── file walker ────────────────────────────────────────────────────────────────

def walk_xlsx(root: Path):
    """Yield all non-temp .xlsx paths under root (recursive)."""
    seen = set()
    for p in sorted(root.rglob("*.xlsx")):
        if not p.name.startswith("~") and p not in seen:
            seen.add(p)
            yield p

# ── main scraper ───────────────────────────────────────────────────────────────

def scrape(root: Path) -> dict:
    all_configs:      list[dict] = []
    all_bindings:     list[dict] = []
    all_inputs_ref:   list[dict] = []
    all_outputs_ref:  list[dict] = []
    all_preferences:  list[dict] = []

    config_id = 0

    for xlsx_path in walk_xlsx(root):
        rel        = xlsx_path.relative_to(root)
        folder     = rel.parts[0] if len(rel.parts) > 1 else ""
        rel_path   = str(rel)
        xlsx_name  = xlsx_path.name

        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        except Exception as e:
            print(f"  SKIP {rel}: {e}", file=sys.stderr)
            continue

        file_mode_count = 0
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws   = wb[sheet_name]
            rows = _load_rows(ws)
            kind = classify_sheet(sheet_name, rows[:6])

            if kind == "config":
                cfgs, bnds = parse_config_sheet(
                    rows, xlsx_name, folder, rel_path, sheet_name, sheet_idx
                )
                # Assign sequential config_ids and link bindings
                block_to_cid: dict[int, int] = {}
                for cfg in cfgs:
                    cfg["config_id"] = config_id
                    block_to_cid[cfg["block_index"]] = config_id
                    config_id += 1
                    file_mode_count += 1
                all_configs.extend(cfgs)

                for bnd in bnds:
                    bi = bnd.pop("_block_index")
                    if bi in block_to_cid:
                        bnd["config_id"] = block_to_cid[bi]
                        all_bindings.append(bnd)

            elif kind == "inputs_ref":
                all_inputs_ref.extend(parse_inputs_ref(rows, xlsx_name))

            elif kind == "outputs_ref":
                all_outputs_ref.extend(parse_outputs_ref(rows, xlsx_name))

            elif kind == "preferences":
                all_preferences.extend(parse_preferences(rows, xlsx_name, sheet_name))

        wb.close()
        print(f"  {rel}  ({file_mode_count} modes)")

    return {
        "configs":      all_configs,
        "bindings":     all_bindings,
        "inputs_ref":   all_inputs_ref,
        "outputs_ref":  all_outputs_ref,
        "preferences":  all_preferences,
    }

# ── database writer ────────────────────────────────────────────────────────────

SCHEMA_SQL = """
CREATE TABLE configs (
    config_id    INTEGER PRIMARY KEY,
    folder       TEXT,
    xlsx_name    TEXT,
    rel_path     TEXT,
    sheet_name   TEXT,
    sheet_index  INTEGER,
    block_index  INTEGER,
    mode_name    TEXT,
    layout_type  TEXT,
    transport    TEXT,
    profile_file TEXT
);

CREATE TABLE bindings (
    binding_id    INTEGER PRIMARY KEY AUTOINCREMENT,
    config_id     INTEGER REFERENCES configs(config_id),
    output_name   TEXT,
    function_type TEXT,
    input_1  TEXT,
    input_2  TEXT,
    input_3  TEXT,
    input_4  TEXT,
    input_5  TEXT,
    input_6  TEXT,
    input_7  TEXT,
    input_8  TEXT
);

CREATE TABLE inputs_ref (
    rowid        INTEGER PRIMARY KEY AUTOINCREMENT,
    input_name   TEXT,
    description  TEXT,
    delay        TEXT,
    source_xlsx  TEXT
);

CREATE TABLE outputs_ref (
    rowid        INTEGER PRIMARY KEY AUTOINCREMENT,
    output_name  TEXT,
    description  TEXT,
    source_xlsx  TEXT
);

CREATE TABLE preferences (
    pref_id      INTEGER PRIMARY KEY AUTOINCREMENT,
    source_xlsx  TEXT,
    sheet_name   TEXT,
    preference   TEXT,
    value        TEXT,
    units        TEXT,
    description  TEXT
);

-- How often each output (control) is bound across all configs
CREATE VIEW output_usage AS
SELECT
    output_name,
    COUNT(DISTINCT config_id) AS config_count,
    COUNT(*)                  AS binding_count
FROM bindings
GROUP BY output_name
ORDER BY config_count DESC, binding_count DESC;

-- How often each input gesture is used
CREATE VIEW input_usage AS
SELECT input_name, COUNT(*) AS usage_count FROM (
    SELECT input_1 AS input_name FROM bindings WHERE input_1 IS NOT NULL UNION ALL
    SELECT input_2 FROM bindings WHERE input_2 IS NOT NULL UNION ALL
    SELECT input_3 FROM bindings WHERE input_3 IS NOT NULL UNION ALL
    SELECT input_4 FROM bindings WHERE input_4 IS NOT NULL UNION ALL
    SELECT input_5 FROM bindings WHERE input_5 IS NOT NULL UNION ALL
    SELECT input_6 FROM bindings WHERE input_6 IS NOT NULL UNION ALL
    SELECT input_7 FROM bindings WHERE input_7 IS NOT NULL UNION ALL
    SELECT input_8 FROM bindings WHERE input_8 IS NOT NULL
)
GROUP BY input_name ORDER BY usage_count DESC;

-- Boolean feature matrix for each config – useful for clustering
CREATE VIEW config_features AS
SELECT
    c.config_id,
    c.folder,
    c.xlsx_name,
    c.sheet_name,
    c.mode_name,
    c.transport,
    -- joysticks
    MAX(CASE WHEN b.output_name LIKE 'left_joy%'  THEN 1 ELSE 0 END) AS has_left_joystick,
    MAX(CASE WHEN b.output_name LIKE 'right_joy%' THEN 1 ELSE 0 END) AS has_right_joystick,
    -- d-pad
    MAX(CASE WHEN b.output_name LIKE 'dpad%'      THEN 1 ELSE 0 END) AS has_dpad,
    -- face buttons (PS naming)
    MAX(CASE WHEN b.output_name = 'x'        THEN 1 ELSE 0 END) AS has_x,
    MAX(CASE WHEN b.output_name = 'circle'   THEN 1 ELSE 0 END) AS has_circle,
    MAX(CASE WHEN b.output_name = 'square'   THEN 1 ELSE 0 END) AS has_square,
    MAX(CASE WHEN b.output_name = 'triangle' THEN 1 ELSE 0 END) AS has_triangle,
    -- shoulder / trigger
    MAX(CASE WHEN b.output_name IN ('left_1','right_1')  THEN 1 ELSE 0 END) AS has_l1_r1,
    MAX(CASE WHEN b.output_name IN ('left_2','right_2')  THEN 1 ELSE 0 END) AS has_l2_r2,
    MAX(CASE WHEN b.output_name IN ('left_3','right_3')  THEN 1 ELSE 0 END) AS has_l3_r3,
    -- meta
    MAX(CASE WHEN b.output_name IN ('start','select','ps3','options') THEN 1 ELSE 0 END) AS has_meta_btns,
    -- keyboard / mouse outputs
    MAX(CASE WHEN b.output_name LIKE 'key_%'   THEN 1 ELSE 0 END) AS has_keyboard,
    MAX(CASE WHEN b.output_name LIKE 'mouse%'  THEN 1 ELSE 0 END) AS has_mouse,
    -- mode switching
    MAX(CASE WHEN b.output_name IN ('increment_mode','decrement_mode') THEN 1 ELSE 0 END) AS has_mode_switch,
    COUNT(DISTINCT b.output_name) AS total_bound_outputs
FROM configs c
LEFT JOIN bindings b ON b.config_id = c.config_id
GROUP BY c.config_id;
"""


def write_db(data: dict, db_path: Path):
    db_path.parent.mkdir(parents=True, exist_ok=True)
    db_path.unlink(missing_ok=True)
    conn = sqlite3.connect(db_path)
    conn.executescript(SCHEMA_SQL)

    def ins(table, rows):
        if not rows:
            return
        cols = list(rows[0].keys())
        ph   = ", ".join("?" for _ in cols)
        col_str = ", ".join(f'"{c}"' for c in cols)
        conn.executemany(
            f'INSERT INTO "{table}" ({col_str}) VALUES ({ph})',
            [[r.get(c) for c in cols] for r in rows]
        )

    ins("configs",     data["configs"])
    ins("bindings",    data["bindings"])
    ins("inputs_ref",  data["inputs_ref"])
    ins("outputs_ref", data["outputs_ref"])
    ins("preferences", data["preferences"])

    conn.commit()
    conn.close()


def write_csvs(data: dict, csv_dir: Path):
    csv_dir.mkdir(parents=True, exist_ok=True)
    for table, rows in data.items():
        if not rows:
            continue
        out = csv_dir / f"qs_{table}.csv"
        cols = list(rows[0].keys())
        with open(out, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=cols)
            w.writeheader()
            w.writerows(rows)
        print(f"  CSV  {out}  ({len(rows):,} rows)")

# ── entry point ────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser(description="Scrape QuadStick .xlsx configs to SQLite")
    ap.add_argument("--root",    default=DEFAULT_ROOT,
                    help="Root folder containing config sub-directories")
    ap.add_argument("--db",      default=str(DEFAULT_DB),
                    help="Output SQLite database path")
    ap.add_argument("--csv-dir", default=str(DEFAULT_CSV_DIR),
                    help="Directory for CSV exports (empty string to skip)")
    args = ap.parse_args()

    root    = Path(args.root)
    db_path = Path(args.db)
    csv_dir = Path(args.csv_dir) if args.csv_dir else None

    if not root.exists():
        sys.exit(f"Root not found: {root}")

    print(f"Scanning: {root}")
    data = scrape(root)

    n_configs  = len(data["configs"])
    n_bindings = len(data["bindings"])
    print(f"\nCollected: {n_configs:,} mode-configs, {n_bindings:,} bindings")
    print(f"           {len(data['inputs_ref']):,} input refs, "
          f"{len(data['outputs_ref']):,} output refs, "
          f"{len(data['preferences']):,} preferences")

    print(f"\nWriting SQLite -> {db_path}")
    write_db(data, db_path)
    print(f"  done  ({db_path.stat().st_size:,} bytes)")

    if csv_dir:
        print(f"\nWriting CSVs -> {csv_dir}")
        write_csvs(data, csv_dir)

    # Quick summary: most-used outputs across all configs
    conn = sqlite3.connect(db_path)
    print("\nTop 20 outputs by config coverage:")
    print(f"  {'output_name':<30} {'configs':>8}  {'bindings':>9}")
    print("  " + "-" * 52)
    for row in conn.execute(
        "SELECT output_name, config_count, binding_count FROM output_usage LIMIT 20"
    ):
        print(f"  {row[0]:<30} {row[1]:>8}  {row[2]:>9}")
    conn.close()

    print("\nDone.")


if __name__ == "__main__":
    main()
